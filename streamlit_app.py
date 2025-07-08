import streamlit as st 
import pandas as pd
import math
import os
import openpyxl
from openpyxl import load_workbook

st.title("使用量と必要本数シミュレーター")

# === ✅ Excelファイルの選択肢（data/フォルダ構成） ===
file_options = {
    "K40": os.path.join("data", "32Rk40.xlsx"),
    "1085G": os.path.join("data", "1085G使用量.xlsx"),
    "E51G-JP": os.path.join("data", "E51G-JP使用量.xlsx")
}

# ファイル選択
selected_file_key = st.sidebar.selectbox("材質選択", list(file_options.keys()))
file_path = file_options[selected_file_key]

try:
    df = pd.read_excel(file_path, engine="openpyxl")

    # 使用量列の単位除去と変換
    df["使用量"] = (
        df["使用量"]
        .astype(str)
        .replace(r"[gG\s]", "", regex=True)
        .replace("", "0")
        .astype(float)
    )

    st.sidebar.header("⚙️ 設定")

    selected_processes = st.sidebar.multiselect(
        "工程を選択", options=df["工程"].unique(), default=list(df["工程"].unique())
    )
    operating_days = st.sidebar.number_input("稼働日数（生産日）", min_value=1, value=20)
    production_units = st.sidebar.number_input("1日/生産台数", min_value=1, value=1100)
    drum_capacity = st.sidebar.number_input("ドラム缶容量 (kg)", min_value=1.0, value=250.0, step=10.0)
    split_days = st.sidebar.number_input("振り分け日数", min_value=1, value=15)
    loss_per_drum = st.sidebar.number_input("交換時エアー抜き量 (kg)", min_value=0.0, max_value=drum_capacity - 1, value=20.0)

    # 実質使用可能容量（ロスを除いた容量）
    usable_capacity = drum_capacity - loss_per_drum

    # フィルタリング
    filtered_df = df[df["工程"].isin(selected_processes)]

    # 工程ごとの1台あたり使用量（g）
    per_unit = filtered_df.groupby("工程")["使用量"].sum().reset_index()
    per_unit.columns = ["工程", "1台あたり使用量（g）"]

    # 総使用量（kg）
    per_unit["総使用量（kg）"] = (
        per_unit["1台あたり使用量（g）"] * production_units * operating_days / 1000
    )

    # 必要ドラム缶数
    per_unit["必要ドラム缶数"] = per_unit["総使用量（kg）"].apply(
        lambda x: math.ceil(x / usable_capacity)
    )

    # 全体集計
    total_required_kg = per_unit["総使用量（kg）"].sum()
    total_drum_count = total_required_kg / usable_capacity
    daily_drum_count = total_drum_count / split_days
    total_loss_kg = per_unit["必要ドラム缶数"].sum() * loss_per_drum

    # 表示用データ整形
    per_unit_display = per_unit.copy()
    per_unit_display["総使用量（kg）"] = per_unit_display["総使用量（kg）"].map(lambda x: f"{x:.1f} kg")
    per_unit_display["必要ドラム缶数"] = per_unit_display["必要ドラム缶数"].astype(str) + " 本"

    # 表示
    st.subheader(f"📋 工程ごとの必要本数（kg）と必要ドラム缶数 [{selected_file_key}]")
    st.dataframe(per_unit_display)

    st.subheader("📌 総使用量の合計と日別振り分け（ドラム缶本数）")
    st.markdown(f"✅ 全工程の必要本数 合計: **{total_drum_count:.1f} 本**")
    st.markdown(f"📅 {split_days}日で振り分けた場合：**1日あたり {daily_drum_count:.1f} 本**")
    st.markdown(f"♻️ ドラム交換による総ロス見込み: **{total_loss_kg:.1f} kg**")

    # ===== 📆 発注スケジュール入力エリア =====
    st.subheader("📆 発注スケジュール（週×曜日）入力")

    days = ["月", "火", "水", "木", "金"]
    weeks = [f"{i+1}週目" for i in range((split_days // 5) + 1)]

    schedule = {}
    total_input = 0

    for week in weeks:
        st.markdown(f"**{week}**")
        cols = st.columns(len(days))
        for i, day in enumerate(days):
            key = f"{week}_{day}"
            val = cols[i].number_input(f"{day}", key=key, min_value=0, step=1, value=0)
            schedule[key] = val
            total_input += val

    st.markdown("---")
    st.markdown(f"🧮 入力した合計本数: **{total_input} 本**")
    st.markdown(f"🔢 自動計算した必要本数: **{math.ceil(total_drum_count)} 本**")

    if total_input != math.ceil(total_drum_count):
        st.warning("⚠️ 入力された本数が必要本数と一致していません。")
    
    else:
        st.success("✅ 入力されたスケジュールと必要本数が一致しています。")


if st.button("✅ 確定してExcelに保存"):
    try:
        # ここに処理を書く（省略せずに全部書く）
        template_path = r"C:\Users\J0134011\OneDrive - Honda\デスクトップ\my_app\my_streamlit_app\calendar_template.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active

        start_col = 3  # C列は index 3

        # 工程×材質のマッピング（Excelの行）
        row_map = {
            ("RR Door", "K40"): 2,
            ("FR Door", "K40"): 3,
            ("RR Door", "E51G-JP"): 2,
            ("FR Door", "E51G-JP"): 3,
            ("D7", "1085G"): 4,
        }

        for process in selected_processes:
            key = (process, selected_file_key)
            if key in row_map:
                row = row_map[key]
                col_index = 0
                for week in weeks:
                    for day in days:
                        cell_value = schedule.get(f"{week}_{day}", 0)
                        ws.cell(row=row, column=start_col + col_index, value=cell_value)
                        col_index += 1

        wb.save(template_path)
        st.success("✅ スケジュールをExcelに保存しました")

    except FileNotFoundError:
        st.error("❌ calendar_template.xlsx が見つかりません。ファイルパスを確認してください。")

    except Exception as e:
        st.error(f"⚠️ 保存中にエラーが発生しました: {e}")
